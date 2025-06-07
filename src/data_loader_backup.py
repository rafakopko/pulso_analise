"""
M√≥dulo para carregamento e processamento de dados Excel
"""

import pandas as pd
import streamlit as st
from pathlib import Path
import openpyxl
from typing import Dict, List, Optional, Tuple
import logging

from config import REQUIRED_COLUMNS, EXCEL_CONFIG, MESSAGES, BASE_EXCEL_PATH

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DataLoader:
    """Classe para carregamento e valida√ß√£o de dados do Excel"""
    def __init__(self):
        self.data: Dict[str, pd.DataFrame] = {}
        self.metadata: Dict = {}
    
    def _validate_path_file(self, file_path: Path) -> Tuple[bool, str]:
        """
        Valida um arquivo Excel local usando Path
        
        Args:
            file_path: Path para o arquivo Excel
            
        Returns:
            Tuple[bool, str]: (√©_v√°lido, mensagem)
        """
        try:
            # Verificar se arquivo existe
            if not file_path.exists():
                return False, f"Arquivo n√£o encontrado: {file_path}"
            
            # Verificar tamanho
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            if file_size_mb > EXCEL_CONFIG["max_file_size"]:
                return False, f"Arquivo muito grande. M√°ximo: {EXCEL_CONFIG['max_file_size']}MB"
            
            # Verificar extens√£o
            file_extension = file_path.suffix.lower()
            if file_extension not in [f".{ext}" for ext in EXCEL_CONFIG["file_types"]]:
                return False, f"Tipo de arquivo inv√°lido. Use: {', '.join(EXCEL_CONFIG['file_types'])}"
            
            # Verificar se √© um arquivo Excel v√°lido
            try:
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
            except Exception as e:
                return False, f"Arquivo Excel corrompido: {str(e)}"
            
            # Verificar abas obrigat√≥rias
            missing_sheets = []
            for required_sheet in EXCEL_CONFIG["required_sheets"]:
                if required_sheet not in sheet_names:
                    missing_sheets.append(required_sheet)
            
            if missing_sheets:
                return False, f"Abas n√£o encontradas: {', '.join(missing_sheets)}"
            
            return True, "Arquivo v√°lido"
              except Exception as e:
            logger.error(f"Erro na valida√ß√£o do arquivo: {str(e)}")
            return False, f"Erro na valida√ß√£o: {str(e)}"
    
    def validate_file(self, file) -> Tuple[bool, str]:
        """
        Valida se o arquivo Excel est√° no formato correto
        
        Args:
            file: Arquivo enviado via streamlit file_uploader ou Path object
            
        Returns:
            Tuple[bool, str]: (√©_v√°lido, mensagem)
        """
        try:
            # Verificar se √© um Path ou arquivo do Streamlit
            if isinstance(file, Path):
                return self._validate_path_file(file)
            
            # Valida√ß√£o para arquivo do Streamlit
            if hasattr(file, 'size') and file.size > EXCEL_CONFIG["max_file_size"] * 1024 * 1024:
                return False, f"Arquivo muito grande. M√°ximo: {EXCEL_CONFIG['max_file_size']}MB"
            
            # Verificar extens√£o
            file_name = file.name if hasattr(file, 'name') else str(file)
            file_extension = Path(file_name).suffix.lower()
            if file_extension not in [f".{ext}" for ext in EXCEL_CONFIG["file_types"]]:
                return False, f"Tipo de arquivo inv√°lido. Use: {', '.join(EXCEL_CONFIG['file_types'])}"
            
            # Verificar se √© um arquivo Excel v√°lido
            try:
                workbook = openpyxl.load_workbook(file, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
            except Exception as e:
                return False, f"Arquivo Excel corrompido: {str(e)}"
            
            # Verificar abas obrigat√≥rias
            missing_sheets = []
            for required_sheet in EXCEL_CONFIG["required_sheets"]:
                if required_sheet not in sheet_names:
                    missing_sheets.append(required_sheet)
            
            if missing_sheets:
                return False, f"Abas n√£o encontradas: {', '.join(missing_sheets)}"
            
            return True, "Arquivo v√°lido"
            
        except Exception as e:
            logger.error(f"Erro na valida√ß√£o do arquivo: {str(e)}")
            return False, f"Erro na valida√ß√£o: {str(e)}"
    
    def load_excel_data(self, file) -> Dict[str, pd.DataFrame]:
        """
        Carrega dados das abas principais do Excel
        
        Args:
            file: Arquivo Excel enviado
            
        Returns:
            Dict[str, pd.DataFrame]: Dicion√°rio com DataFrames das abas
        """
        try:
            # Validar arquivo
            is_valid, message = self.validate_file(file)
            if not is_valid:
                st.error(f"‚ùå {message}")
                return {}
            
            # Carregar abas principais
            data = {}
            
            with st.spinner("üîÑ Carregando dados..."):
                # Aba principal - dados di√°rios
                try:
                    df_diaria = pd.read_excel(
                        file, 
                        sheet_name="pulso_consulta_diaria",
                        header=1,  # Skip description row
                        engine='openpyxl'
                    )
                    data["pulso_consulta_diaria"] = df_diaria
                    logger.info(f"Carregada aba pulso_consulta_diaria: {df_diaria.shape}")
                    
                except Exception as e:
                    st.error(f"‚ùå Erro ao carregar aba 'pulso_consulta_diaria': {str(e)}")
                    return {}
                
                # Aba com c√°lculos de lacunas
                try:
                    df_cluster = pd.read_excel(
                        file,
                        sheet_name="pulso_consulta_diaria_cluster_a", 
                        header=1,  # Skip description row
                        engine='openpyxl'
                    )
                    data["pulso_consulta_diaria_cluster_a"] = df_cluster
                    logger.info(f"Carregada aba cluster: {df_cluster.shape}")
                    
                except Exception as e:
                    st.error(f"‚ùå Erro ao carregar aba 'pulso_consulta_diaria_cluster_a': {str(e)}")
                    return {}
            
            # Validar colunas obrigat√≥rias
            validation_success = True
            for sheet_name, df in data.items():
                missing_cols = self._validate_columns(df, sheet_name)
                if missing_cols:
                    st.warning(f"‚ö†Ô∏è Colunas n√£o encontradas em '{sheet_name}': {', '.join(missing_cols)}")
                    validation_success = False
            
            if not validation_success:
                st.warning("‚ö†Ô∏è Algumas colunas obrigat√≥rias n√£o foram encontradas. A aplica√ß√£o pode n√£o funcionar corretamente.")
            
            # Processar dados b√°sicos
            data = self._preprocess_data(data)
            
            # Salvar metadados
            self.metadata = {
                "file_name": file.name,
                "file_size": file.size,
                "upload_time": pd.Timestamp.now(),
                "total_records": sum(len(df) for df in data.values()),
                "sheets_loaded": list(data.keys())
            }
            
            self.data = data
            st.success(MESSAGES["upload_success"])
            
            return data
            
        except Exception as e:
            logger.error(f"Erro no carregamento dos dados: {str(e)}")
            st.error(f"‚ùå Erro no carregamento: {str(e)}")
            return {}
    
    def _validate_columns(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """
        Valida se as colunas obrigat√≥rias existem no DataFrame
        
        Args:
            df: DataFrame a ser validado
            sheet_name: Nome da aba/sheet
            
        Returns:
            List[str]: Lista de colunas faltantes
        """
        if sheet_name not in REQUIRED_COLUMNS:
            return []
        
        required_cols = REQUIRED_COLUMNS[sheet_name]
        existing_cols = df.columns.tolist()
        missing_cols = [col for col in required_cols if col not in existing_cols]
        
        return missing_cols
    
    def _preprocess_data(self, data: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """
        Pr√©-processa os dados carregados
        
        Args:
            data: Dicion√°rio com DataFrames
            
        Returns:
            Dict[str, pd.DataFrame]: Dados processados
        """
        processed_data = {}
        
        for sheet_name, df in data.items():
            df_processed = df.copy()
            
            # Convers√µes comuns
            if 'datavenda' in df_processed.columns:
                df_processed['datavenda'] = pd.to_datetime(df_processed['datavenda'], errors='coerce')
            
            # Limpar nomes de colunas
            df_processed.columns = df_processed.columns.str.strip()
            
            # Substituir valores nulos em colunas num√©ricas por 0
            numeric_cols = df_processed.select_dtypes(include=['number']).columns
            df_processed[numeric_cols] = df_processed[numeric_cols].fillna(0)
            
            # Substituir valores nulos em colunas de texto por string vazia
            text_cols = df_processed.select_dtypes(include=['object']).columns
            df_processed[text_cols] = df_processed[text_cols].fillna('')
            
            processed_data[sheet_name] = df_processed
            
            logger.info(f"Pr√©-processamento de '{sheet_name}' conclu√≠do: {df_processed.shape}")
        
        return processed_data
    
    def get_data_summary(self) -> Dict:
        """
        Retorna resumo dos dados carregados
        
        Returns:
            Dict: Resumo dos dados
        """
        if not self.data:
            return {}
        
        summary = {
            "metadata": self.metadata,
            "sheets": {}
        }
        
        for sheet_name, df in self.data.items():
            summary["sheets"][sheet_name] = {
                "rows": len(df),
                "columns": len(df.columns),
                "columns_list": df.columns.tolist(),
                "memory_usage": df.memory_usage(deep=True).sum()
            }
        
        return summary
    
    def load_default_base(self) -> Dict[str, pd.DataFrame]:
        """
        Carrega a base padr√£o do diret√≥rio Base/
        
        Returns:
            Dict[str, pd.DataFrame]: Dados da base padr√£o ou dict vazio se erro
        """
        try:
            if not BASE_EXCEL_PATH.exists():
                logger.warning(f"Arquivo base n√£o encontrado: {BASE_EXCEL_PATH}")
                return {}
            
            logger.info(f"Carregando base padr√£o: {BASE_EXCEL_PATH}")
            
            # Carregar abas principais
            data = {}
              # Aba principal - dados di√°rios
            try:
                df_diaria = pd.read_excel(
                    BASE_EXCEL_PATH, 
                    sheet_name="pulso_consulta_diaria",
                    header=1,  # Skip description row
                    engine='openpyxl'
                )
                data["pulso_consulta_diaria"] = df_diaria
                logger.info(f"Carregada aba pulso_consulta_diaria da base padr√£o: {df_diaria.shape}")
                
            except Exception as e:
                logger.error(f"Erro ao carregar aba 'pulso_consulta_diaria' da base padr√£o: {str(e)}")                return {}
            # Aba com c√°lculos de lacunas
            try:
                df_cluster = pd.read_excel(
                    BASE_EXCEL_PATH,
                    sheet_name="pulso_consulta_diaria_cluster_a", 
                    header=1,  # Skip description row
                    engine='openpyxl'
                )
                data["pulso_consulta_diaria_cluster_a"] = df_cluster
                logger.info(f"Carregada aba cluster da base padr√£o: {df_cluster.shape}")
                
            except Exception as e:
                logger.error(f"Erro ao carregar aba 'pulso_consulta_diaria_cluster_a' da base padr√£o: {str(e)}")
                return {}
        
            # Validar colunas obrigat√≥rias
            validation_success = True
            for sheet_name, df in data.items():
                missing_cols = self._validate_columns(df, sheet_name)
                if missing_cols:
                    logger.warning(f"Colunas n√£o encontradas em '{sheet_name}': {', '.join(missing_cols)}")
                    validation_success = False
            
            if not validation_success:
                logger.warning("Algumas colunas obrigat√≥rias n√£o foram encontradas na base padr√£o.")
            
            # Processar dados b√°sicos
            data = self._preprocess_data(data)
            
            # Salvar metadados
            file_stat = BASE_EXCEL_PATH.stat()
            self.metadata = {
                "file_name": BASE_EXCEL_PATH.name,
                "file_size": file_stat.st_size,
                "upload_time": pd.Timestamp.now(),
                "total_records": sum(len(df) for df in data.values()),
                "sheets_loaded": list(data.keys()),
                "source": "base_padrao"
            }
            
            self.data = data
            logger.info("Base padr√£o carregada com sucesso!")
            
            return data
            
        except Exception as e:
            logger.error(f"Erro ao carregar base padr√£o: {str(e)}")
            return {}

def check_default_base_exists() -> bool:
    """
    Verifica se a base padr√£o existe
    
    Returns:
        bool: True se a base padr√£o existe
    """
    return BASE_EXCEL_PATH.exists()


def load_default_base_if_exists() -> Optional[Dict[str, pd.DataFrame]]:
    """
    Carrega a base padr√£o se ela existir
    
    Returns:
        Optional[Dict[str, pd.DataFrame]]: Dados carregados ou None
    """
    if check_default_base_exists():
        loader = DataLoader()
        return loader.load_default_base()
    return None


def load_sample_data() -> Dict[str, pd.DataFrame]:
    """
    Carrega dados de exemplo para demonstra√ß√£o
    
    Returns:
        Dict[str, pd.DataFrame]: Dados de exemplo
    """
    # Criar dados fict√≠cios para demonstra√ß√£o
    sample_data = {
        "pulso_consulta_diaria": pd.DataFrame({
            "NomeLoja": [f"Loja {i:03d}" for i in range(1, 21)],
            "codigo_franquia": range(1, 21),
            "NumeroGR": [1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3, 4, 4, 4, 4, 4],
            "datavenda": pd.date_range("2025-01-01", periods=20, freq="D"),
            "receita_liquida": [10000 + i * 1000 for i in range(20)],
            "qtd_cupom": [100 + i * 10 for i in range(20)],
            "qtd_item": [200 + i * 20 for i in range(20)],
            "Mediana_Semana_RL": [15000] * 20,
            "Mediana_Semana_cupom": [150] * 20
        }),        "pulso_consulta_diaria_cluster_a": pd.DataFrame({
            "NomeLoja": [f"Loja {i:03d}" for i in range(1, 21)],
            "grupo_comparavel": ["1-0", "1-0", "2-0", "2-0", "3-0"] * 4,
            "LacunaRL": [-5000, -3000, -1000, 2000, 4000] * 4,
            "LacunaCupom": [-50, -30, -10, 20, 40] * 4,
            "LacunaBM": [-100, -60, -20, 40, 80] * 4,
            "LacunaPM": [-10, -6, -2, 4, 8] * 4,
            "LacunaProd": [-5, -3, -1, 2, 4] * 4
        })
    }
    
    return sample_data
